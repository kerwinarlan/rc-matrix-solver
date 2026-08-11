"""Thin Excel read/write layer with the engine behind one interface.

Primary engine: openpyxl (headless .xlsx, no Excel install needed).
Fallback engine: xlwings (live Excel, documented stub - see the architecture
doc for the design). run.py depends only on the WorkbookIO interface, so the
engine is swappable without touching the orchestrator.

Self-check: ``python3 -m bridge.excel_io`` builds a template, reads and writes
it back, and asserts the round-trip.
"""

from pathlib import Path
from typing import Dict, List, Optional, Protocol, Union

from .workbook_layout import MAX_ROWS, SHEETS, TABLES, Table, column_letter

try:
    import openpyxl
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:  # pragma: no cover - only reachable without openpyxl installed
    openpyxl = None  # type: ignore[assignment]

InputValue = Union[int, float]
InputData = Dict[str, List[InputValue]]
OutputData = Dict[str, List[InputValue]]

_OPENPYXL_MISSING = (
    "openpyxl is not installed; run `python3 -m pip install openpyxl` "
    "(or switch engines - see docs/excel-bridge-architecture.md)"
)


class WorkbookIO(Protocol):
    """Engine interface. The orchestrator depends only on this protocol."""

    def read_inputs(self) -> InputData:
        """Return input tables keyed by named-range name, e.g. ``NODE_X`` -> [0.0, 5.0]."""
        ...

    def write_outputs(self, outputs: OutputData) -> None:
        """Write output tables keyed by named-range name, e.g. ``OUT_DISP_UX`` -> [...].

        Unknown names raise ValueError. Rows are written in dict order; shorter
        lists are padded with blanks. Stale cells in the written band are cleared.
        """
        ...


class OpenpyxlWorkbookIO:
    """Reads/writes a .xlsx file headlessly. Requires ``pip install openpyxl``."""

    def __init__(self, path: Union[str, Path]) -> None:
        if openpyxl is None:
            raise RuntimeError(_OPENPYXL_MISSING)
        self.path = Path(path)

    def read_inputs(self) -> InputData:
        """Scan each input table until its first blank row; return flat column dicts."""
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        data: InputData = {}
        for table in TABLES:
            if table.direction != "in":
                continue
            rows = _read_table_rows(wb[table.sheet], table)
            for index, column in enumerate(table.columns):
                data[column.name] = [_coerce(row[index], column.dtype) for row in rows]
        return data

    def write_outputs(self, outputs: OutputData) -> None:
        """Write the output tables. See WorkbookIO.write_outputs for semantics."""
        known = {column.name: (table, index) for table in TABLES
                 for index, column in enumerate(table.columns)}
        unknown = [name for name in outputs if name not in known]
        if unknown:
            raise ValueError(f"unknown output range names: {unknown}")

        wb = openpyxl.load_workbook(self.path)
        for table in TABLES:
            if table.direction != "out":
                continue
            hits = [(index, column) for index, column in enumerate(table.columns)
                    if column.name in outputs]
            if not hits:
                continue
            ws = wb[table.sheet]
            count = max(len(outputs[column.name]) for _, column in hits)
            for index, column in hits:
                values = list(outputs[column.name])
                values.extend([None] * (count - len(values)))
                for offset, value in enumerate(values):
                    ws.cell(table.data_row + offset, table.start_col + index, value)
            _clear_stale(ws, table)
        wb.save(self.path)

    def __repr__(self) -> str:  # pragma: no cover - debugging aid
        return f"OpenpyxlWorkbookIO({str(self.path)!r})"


class XlwingsWorkbookIO:
    """Live-Excel engine (xlwings). Documented stub - not implemented.

    Design (see docs/excel-bridge-architecture.md): open a workbook via
    ``xlwings.Book``, read/write the same named ranges from workbook_layout.py,
    and provide ``read_inputs`` / ``write_outputs`` like OpenpyxlWorkbookIO.
    Requires Excel installed on the machine and ``pip install xlwings``.
    """

    def __init__(self, path: Union[str, Path]) -> None:
        raise NotImplementedError(
            "xlwings engine is a documented stub; use engine='openpyxl' for now "
            "- see docs/excel-bridge-architecture.md"
        )


def open_workbook(path: Union[str, Path], engine: str = "openpyxl") -> WorkbookIO:
    """Factory over the WorkbookIO interface. engine: "openpyxl" (default) or "xlwings"."""
    if engine == "openpyxl":
        return OpenpyxlWorkbookIO(path)
    if engine == "xlwings":
        raise NotImplementedError(
            "xlwings engine is a documented stub; see docs/excel-bridge-architecture.md"
        )
    raise ValueError(f"unknown engine {engine!r}: choose 'openpyxl' or 'xlwings'")


def build_template(path: Union[str, Path]) -> Path:
    """Generate a starter .xlsx from workbook_layout.py: headers, count cells,
    bold header row, freeze panes, and all named ranges (rows up to MAX_ROWS)."""
    if openpyxl is None:
        raise RuntimeError(_OPENPYXL_MISSING)
    path = Path(path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in SHEETS:
        wb.create_sheet(sheet)
    for table in TABLES:
        ws = wb[table.sheet]
        for index, column in enumerate(table.columns):
            cell = ws.cell(table.start_row, table.start_col + index, column.header)
            cell.font = openpyxl.styles.Font(bold=True)
        if table.count_name:
            _write_count(ws, table)
        ws.freeze_panes = ws.cell(table.data_row, table.start_col)
    _define_names(wb)
    wb.save(path)
    return path


# --- internals -----------------------------------------------------------


def _read_table_rows(ws, table: Table) -> List[List[InputValue]]:
    """Read data rows until the first fully blank row in the table band."""
    rows: List[List[InputValue]] = []
    band = ws.iter_rows(
        min_row=table.data_row,
        max_row=table.data_row + MAX_ROWS,
        min_col=table.start_col,
        max_col=table.start_col + len(table.columns) - 1,
    )
    for row in band:
        values = [cell.value for cell in row]
        if all(value is None for value in values):
            break
        rows.append(values)
    return rows


def _coerce(value: InputValue, dtype: str) -> InputValue:
    """id -> int, number -> float. None never reaches here (blank rows stop the scan)."""
    return int(value) if dtype == "id" else float(value)


def _write_count(ws, table: Table) -> None:
    """Write the count label and a 0 value; user updates the value in Excel."""
    value_cell = table.count_cell
    ws[value_cell] = 0
    label_col = chr(ord(_letters(value_cell)) - 1)  # one column left of the value
    ws[f"{label_col}{_rows(value_cell)}"] = table.count_name


def _clear_stale(ws, table: Table) -> None:
    """Blank the whole output band before writing, so a shorter run never
    leaves previous results behind (None-valued cells are not saved to disk)."""
    for index in range(len(table.columns)):
        for row in range(table.data_row, table.data_row + MAX_ROWS):
            ws.cell(row, table.start_col + index, None)


def _define_names(wb) -> None:
    """Register every column and count name as a workbook-level named range."""
    for table in TABLES:
        for index, column in enumerate(table.columns):
            letter = column_letter(table.start_col + index)
            ref = f"'{table.sheet}'!${letter}${table.data_row}:${letter}${MAX_ROWS}"
            wb.defined_names.add(DefinedName(column.name, attr_text=ref))
        if table.count_name:
            ref = f"'{table.sheet}'!${table.count_cell}"
            wb.defined_names.add(DefinedName(table.count_name, attr_text=ref))


def _letters(cell: str) -> str:
    return "".join(ch for ch in cell if ch.isalpha())


def _rows(cell: str) -> str:
    return "".join(ch for ch in cell if ch.isdigit())


if __name__ == "__main__":  # self-check: one runnable round-trip
    import os
    import tempfile

    assert openpyxl is not None, "self-check requires openpyxl"
    path = Path(tempfile.mktemp(suffix=".xlsx"))
    try:
        build_template(path)
        io = OpenpyxlWorkbookIO(path)

        inputs = io.read_inputs()
        expected_inputs = {column.name for table in TABLES if table.direction == "in"
                           for column in table.columns}
        assert set(inputs) == expected_inputs, "input names mismatch"
        assert all(values == [] for values in inputs.values()), "blank template must read empty"

        outputs: OutputData = {
            "OUT_DISP_NODE_ID": [1, 2], "OUT_DISP_UX": [0.001, -0.002],
            "OUT_DISP_UY": [0.0, 0.0005], "OUT_DISP_RZ": [0.0, 0.0001],
            "OUT_REAC_NODE_ID": [1, 2], "OUT_REAC_FX": [100.0, -100.0],
            "OUT_REAC_FY": [50.0, -50.0], "OUT_REAC_MZ": [0.0, 0.0],
            "OUT_MF_MEMBER_ID": [1], "OUT_MF_AXIAL": [-25000.0], "OUT_MF_SHEAR": [3000.0],
            "OUT_MF_M_I": [0.0], "OUT_MF_M_J": [-4500.0],
            "OUT_DES_MEMBER_ID": [1], "OUT_DES_AS_REQ": [420.0], "OUT_DES_AS_PROV": [452.0],
            "OUT_DES_STIRRUP": [200.0],
        }
        io.write_outputs(outputs)
        wb = openpyxl.load_workbook(path)
        assert wb["Outputs-Displacements"]["B2"].value == 0.001, "ux not written"
        assert wb["Outputs-Displacements"]["C3"].value == 0.0005, "uy row 2 not written"
        assert wb["Outputs-MemberForces"]["A2"].value == 1, "member id not written"
        for name in ("NODE_X", "MEMBER_I", "OUT_DISP_UX", "NODE_COUNT", "LOAD_U_COUNT"):
            assert name in wb.defined_names, f"missing named range {name}"

        wb = openpyxl.load_workbook(path)
        ws = wb["Inputs-Node"]
        ws["A2"], ws["B2"], ws["C2"] = 1, 0.0, 5.0
        ws["D2"], ws["E2"], ws["F2"] = 1, 1, 0
        wb.save(path)
        inputs = OpenpyxlWorkbookIO(path).read_inputs()
        assert inputs["NODE_ID"] == [1] and inputs["NODE_X"] == [0.0]
        assert inputs["NODE_SUP_UX"] == [1.0] and inputs["NODE_SUP_RZ"] == [0.0]

        names = [column.name for table in TABLES for column in table.columns]
        assert len(names) == len(set(names)), "duplicate column names in layout"
        print("bridge.excel_io self-check OK")
    finally:
        os.remove(path)
